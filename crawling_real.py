from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time
from openpyxl import Workbook
from openpyxl.styles import Alignment
from pytube import YouTube
import datetime

SCROLL_PAUSE_TIME = 1.5

keywords = ['fire noodle', '불닭', '불닭 챌린지', 'fire noodle challenge', 'korean spicy noodle', '불닭볶음면', '핵붉닭', 'nuclear fire noodle', 'noodle challenge']

for keyword in keywords:
    url = f"https://www.youtube.com/results?search_query={keyword}"
    service =Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)
    driver.get(url)
    time.sleep(3)

    last_height = driver.execute_script("return document.documentElement.scrollHeight")

    while True:
        driver.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")

        time.sleep(SCROLL_PAUSE_TIME)

        new_height = driver.execute_script("return document.documentElement.scrollHeight;")

        if new_height == last_height:
            break
        last_height = new_height

    try:
        wb = load_workbook("/home/lafesta/Desktop/Methodology/dataset/spicynoodle/"+keyword+".xlsx")
        check_new_file = 0
    except:
        wb = Workbook()
        check_new_file = 1

    if check_new_file ==1:
        while len(wb.sheetnames) > 1:
            wb.remove(wb.sheetnames[len(wb.sheetnames)-1])
        ws = wb.active
        ws.title = keyword+time.strftime("(%Y-%m-%d %H-%M-%S)")
    else:
        ws = wb.create_sheet(keyword+time.strftime("(%Y-%m-%d %H-%M-%S)"))
    
    ws.append(['title','url', 'rating_view', 'published_date'])

    titles = driver.find_elements(By.CSS_SELECTOR, "#dismissible.style-scope.ytd-video-renderer")

    for title in titles:
        main_title = title.find_element(By.CSS_SELECTOR, "#video-title").get_property("title")
        tube_url = title.find_element(By.CSS_SELECTOR, "#video-title").get_property("href")
        ws.append([main_title, tube_url])

    maxrow = ws.max_row
    count_row = maxrow - 1
    k = 1

    for i in range(2, maxrow+1):
        url = ws.cell(row=i, column=2).value
        tube=YouTube(url)
        view = tube.views
        update_dates = str(tube.publish_date)
        update_date = update_dates.split(" ")
        
        ws.cell(row=i,column=3).value = view
        ws.cell(row=i, column=3).number_format = "#,##0"
        ws.cell(row=i, column=4).value = update_date[0]

        print(f"총 {count_row}개 중 {k}번째 완료")
        k+=1
    
    wb.save("/home/lafesta/Desktop/Methodology/dataset/spicynoodle/"+keyword+".xlsx")
    wb.close()